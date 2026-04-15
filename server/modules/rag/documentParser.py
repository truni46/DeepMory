# server/modules/rag/documentParser.py
from __future__ import annotations

import os
from dataclasses import dataclass
from typing import List, Optional

from config.logger import logger


@dataclass
class ParsedPage:
    text: str
    pageNumber: Optional[int]


class PdfParser:
    def parse(self, filePath: str) -> List[ParsedPage]:
        try:
            import pymupdf4llm
            chunks = pymupdf4llm.to_markdown(filePath, page_chunks=True)
            pages = []
            for chunk in chunks:
                text = chunk.get("text", "")
                pageNum = chunk.get("metadata", {}).get("page", 0) + 1
                if text.strip():
                    pages.append(ParsedPage(text=text, pageNumber=pageNum))
            return pages
        except Exception as e:
            logger.error(f"PdfParser.parse failed for '{filePath}': {e}")
            return []


class DocxParser:
    def parse(self, filePath: str) -> List[ParsedPage]:
        # Tier 1: python-docx (handles .docx and OOXML-based .doc)
        try:
            import docx
            doc = docx.Document(filePath)
            text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            if text.strip():
                return [ParsedPage(text=text, pageNumber=None)]
        except Exception as e:
            logger.warning(f"DocxParser: python-docx failed for '{filePath}': {e}")

        # Tier 2: win32com Word automation (handles binary .doc on Windows)
        try:
            import win32com.client
            import pythoncom
            pythoncom.CoInitialize()
            word = win32com.client.Dispatch("Word.Application")
            word.Visible = False
            try:
                absPath = os.path.abspath(filePath)
                doc = word.Documents.Open(absPath, ReadOnly=True)
                text = doc.Content.Text
                doc.Close(False)
            finally:
                word.Quit()
                pythoncom.CoUninitialize()
            if text.strip():
                return [ParsedPage(text=text.strip(), pageNumber=None)]
        except Exception as e:
            logger.warning(f"DocxParser: win32com failed for '{filePath}': {e}")

        # Tier 3: RTF detection + basic strip
        try:
            with open(filePath, "rb") as f:
                header = f.read(6)
            if header.startswith(b"{\\rtf"):
                import re
                with open(filePath, "r", encoding="utf-8", errors="ignore") as f:
                    raw = f.read()
                text = re.sub(r"\\[a-z]+\d*\s?", " ", raw)
                text = re.sub(r"[{}\\]", "", text)
                text = re.sub(r"\s+", " ", text).strip()
                if text:
                    return [ParsedPage(text=text, pageNumber=None)]
        except Exception as e:
            logger.warning(f"DocxParser: RTF fallback failed for '{filePath}': {e}")

        logger.warning(f"DocxParser: all methods failed for '{filePath}', falling back to raw text")
        return TextParser().parse(filePath)


class XlsxParser:
    def parse(self, filePath: str) -> List[ParsedPage]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filePath, read_only=True, data_only=True)
            pages = []
            for sheetName in wb.sheetnames:
                ws = wb[sheetName]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    rows.append(" | ".join(cells))
                text = "\n".join(r for r in rows if r.strip())
                if text.strip():
                    pages.append(ParsedPage(text=text, pageNumber=None))
            return pages
        except Exception as e:
            logger.warning(f"XlsxParser.parse failed for '{filePath}': {e} — falling back to text read")
            return TextParser().parse(filePath)


class TextParser:
    def parse(self, filePath: str) -> List[ParsedPage]:
        try:
            with open(filePath, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()
            return [ParsedPage(text=text, pageNumber=None)] if text.strip() else []
        except Exception as e:
            logger.error(f"TextParser.parse failed for '{filePath}': {e}")
            return []


class DocumentParserService:
    def __init__(self):
        self._parsers = {
            ".pdf": PdfParser(),
            ".docx": DocxParser(),
            ".doc": DocxParser(),
            ".xlsx": XlsxParser(),
            ".xls": XlsxParser(),
        }
        self._default = TextParser()

    def parse(self, filePath: str) -> List[ParsedPage]:
        ext = os.path.splitext(filePath)[1].lower()
        parser = self._parsers.get(ext, self._default)
        return parser.parse(filePath)


documentParserService = DocumentParserService()
